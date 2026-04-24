"""Trade journal logger — appends trades to Trading_Journal.xlsx.

Writes to columns B-I, K, N-Q. Leaves formula columns A, J, L, M, R
intact so Excel auto-calculates PnL metrics.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from config import JOURNAL_FILE

logger = logging.getLogger("crypto_bot.journal")

# Pending entries that failed due to file lock
_pending_entries: list = []


def _find_next_empty_row(ws) -> int:
    """Find the first empty row by scanning column B (Date Opened)."""
    row = 2
    while ws.cell(row=row, column=2).value is not None:
        row += 1
    return row


def _copy_formulas(ws, source_row: int, target_row: int) -> None:
    """Copy formula cells from source_row to target_row.

    Formula columns: A(1), J(10), L(12), M(13), R(18)
    """
    formula_cols = [1, 10, 12, 13, 18]
    for col in formula_cols:
        src_cell = ws.cell(row=source_row, column=col)
        val = src_cell.value
        if val and isinstance(val, str) and val.startswith("="):
            # Replace row references: "2" -> target_row
            new_formula = val.replace(str(source_row), str(target_row))
            ws.cell(row=target_row, column=col).value = new_formula


def log_trade(
    symbol: str,
    direction: str,
    entry_price: float,
    exit_price: Optional[float],
    quantity: float,
    leverage: int = 10,
    fees: float = 0.0,
    strategy: str = "",
    entry_reason: str = "",
    exit_reason: str = "",
    notes: str = "",
    date_opened: Optional[datetime] = None,
    date_closed: Optional[datetime] = None,
) -> bool:
    """Append a trade record to the Trading Journal.

    Returns True if successful, False if file was locked (queued for retry).
    """
    entry = {
        "symbol": symbol,
        "direction": direction,
        "entry_price": entry_price,
        "exit_price": exit_price,
        "quantity": quantity,
        "leverage": leverage,
        "fees": fees,
        "strategy": strategy,
        "entry_reason": entry_reason,
        "exit_reason": exit_reason,
        "notes": notes,
        "date_opened": date_opened or datetime.now(),
        "date_closed": date_closed or datetime.now(),
    }

    try:
        return _write_entry(entry)
    except PermissionError:
        logger.warning("Journal file is locked (open in Excel?). Queuing entry.")
        _pending_entries.append(entry)
        return False
    except Exception as e:
        logger.error("Failed to write journal entry: %s", e)
        _pending_entries.append(entry)
        return False


def _write_entry(entry: dict) -> bool:
    """Write a single trade entry to the Excel file."""
    wb = load_workbook(str(JOURNAL_FILE))
    ws = wb["Trade Log"]

    row = _find_next_empty_row(ws)

    # Copy formulas from row 2 (template row)
    _copy_formulas(ws, source_row=2, target_row=row)

    # Write data columns
    ws.cell(row=row, column=2).value = entry["date_opened"]      # B: Date Opened
    ws.cell(row=row, column=3).value = entry["date_closed"]      # C: Date Closed
    ws.cell(row=row, column=4).value = entry["symbol"]           # D: Symbol
    ws.cell(row=row, column=5).value = entry["direction"]        # E: Direction
    ws.cell(row=row, column=6).value = entry["entry_price"]      # F: Entry Price
    ws.cell(row=row, column=7).value = entry["exit_price"]       # G: Exit Price
    ws.cell(row=row, column=8).value = entry["quantity"]         # H: Quantity
    ws.cell(row=row, column=9).value = entry["leverage"]         # I: Leverage
    ws.cell(row=row, column=11).value = entry["fees"]            # K: Fees
    ws.cell(row=row, column=14).value = entry["strategy"]        # N: Strategy
    ws.cell(row=row, column=15).value = entry["entry_reason"]    # O: Entry Reason
    ws.cell(row=row, column=16).value = entry["exit_reason"]     # P: Exit Reason
    ws.cell(row=row, column=17).value = entry["notes"]           # Q: Notes

    wb.save(str(JOURNAL_FILE))
    wb.close()

    logger.info("Logged trade: %s %s %s @ %.4f -> %.4f",
                entry["direction"], entry["symbol"], entry["strategy"],
                entry["entry_price"], entry.get("exit_price", 0))
    return True


def flush_pending() -> int:
    """Retry writing any pending entries. Returns count of successfully written."""
    if not _pending_entries:
        return 0

    written = 0
    remaining = []
    for entry in _pending_entries:
        try:
            _write_entry(entry)
            written += 1
        except Exception:
            remaining.append(entry)

    _pending_entries.clear()
    _pending_entries.extend(remaining)

    if written:
        logger.info("Flushed %d pending journal entries (%d still queued)",
                     written, len(remaining))
    return written
